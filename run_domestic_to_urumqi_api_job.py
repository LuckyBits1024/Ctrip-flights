import argparse
import os
import time
from datetime import datetime as dt, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill

import ctrip_international_api as city_api
import ctrip_lowest_price_api as lowest_price_api


DEFAULT_DEPARTURE_CITIES = [
    "北京",
    "上海",
    "广州",
    "深圳",
    "成都",
    "重庆",
    "西安",
    "杭州",
    "南京",
    "武汉",
    "长沙",
    "郑州",
    "济南",
    "青岛",
    "厦门",
    "福州",
    "天津",
    "沈阳",
    "大连",
    "哈尔滨",
    "长春",
    "呼和浩特",
    "太原",
    "石家庄",
    "兰州",
    "西宁",
    "银川",
    "拉萨",
    "昆明",
    "贵阳",
    "南宁",
    "海口",
    "三亚",
    "宁波",
    "温州",
    "合肥",
    "南昌",
    "泉州",
    "烟台",
    "威海",
    "珠海",
    "汕头",
    "湛江",
    "南通",
    "徐州",
    "扬州",
    "常州",
    "无锡",
    "台州",
    "义乌",
    "绵阳",
    "宜宾",
    "泸州",
    "丽江",
    "大理",
    "西双版纳",
    "桂林",
    "柳州",
    "北海",
    "张家界",
    "宜昌",
    "襄阳",
    "洛阳",
    "南阳",
    "鄂尔多斯",
    "包头",
    "赤峰",
    "通辽",
    "锡林浩特",
    "乌兰浩特",
    "满洲里",
    "喀什",
    "库尔勒",
    "阿克苏",
    "伊宁",
    "克拉玛依",
    "和田",
    "阿勒泰",
    "哈密",
    "塔城",
    "博乐",
    "库车",
    "莎车",
]

JIANGZHEHUWAN_AIRPORT_CITIES = [
    "上海",
    "南京",
    "无锡",
    "常州",
    "南通",
    "扬州",
    "徐州",
    "盐城",
    "连云港",
    "淮安",
    "杭州",
    "宁波",
    "温州",
    "台州",
    "义乌",
    "舟山",
    "衢州",
    "合肥",
    "黄山",
    "阜阳",
    "安庆",
    "池州",
    "芜湖",
]

XINJIANG_LOCAL_CITIES = {
    "喀什",
    "库尔勒",
    "阿克苏",
    "伊宁",
    "克拉玛依",
    "和田",
    "阿勒泰",
    "哈密",
    "塔城",
    "博乐",
    "库车",
    "莎车",
}

DESTINATION_CITY = "乌鲁木齐"
DESTINATION_CODE = "URC"
LOG_FILE = os.path.join(".codex", "domestic_to_urumqi_api_job.log")


def now_text():
    return dt.now().strftime("%Y-%m-%d_%H-%M-%S")


def log(message):
    line = f"{now_text()} {message}"
    print(line, flush=True)
    os.makedirs(".codex", exist_ok=True)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def result_file_path(run_day, depart_date):
    files_dir = os.path.join(
        os.getcwd(),
        "results",
        "domestic_to_urumqi",
        run_day,
    )
    os.makedirs(files_dir, exist_ok=True)
    return os.path.join(files_dir, f"全国各地-乌鲁木齐_{depart_date}.csv")


def round_trip_result_file_path(run_day, month, stay_days):
    files_dir = os.path.join(
        os.getcwd(),
        "results",
        "domestic_to_urumqi",
        run_day,
    )
    os.makedirs(files_dir, exist_ok=True)
    return os.path.join(files_dir, f"全国各地-乌鲁木齐往返_{month}_{stay_days}天.csv")


def open_jaw_result_file_path(run_day, month, stay_days):
    files_dir = os.path.join(
        os.getcwd(),
        "results",
        "domestic_to_urumqi",
        run_day,
    )
    os.makedirs(files_dir, exist_ok=True)
    return os.path.join(files_dir, f"全国各地-乌鲁木齐异地返程_{month}_{stay_days}天.csv")


def xlsx_path_for(csv_path):
    return os.path.splitext(csv_path)[0] + ".xlsx"


def month_dates(month):
    start = dt.strptime(f"{month}-01", "%Y-%m-%d")
    dates = []
    current = start
    while current.month == start.month:
        dates.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)
    return dates


def filter_cities(cities, include_xinjiang_local=False):
    selected = []
    for city in cities:
        if city == DESTINATION_CITY:
            continue
        if not include_xinjiang_local and city in XINJIANG_LOCAL_CITIES:
            continue
        selected.append(city)
    return selected


def write_table_outputs(frame, csv_path, highlight_missing=False):
    frame.to_csv(csv_path, encoding="UTF-8", index=False)
    xlsx_path = xlsx_path_for(csv_path)
    frame.to_excel(xlsx_path, index=False, engine="openpyxl")

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    if highlight_missing and sheet.max_row > 1:
        missing_fill = PatternFill("solid", fgColor="FFC7CE")
        sheet.conditional_formatting.add(
            f"A2:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}",
            FormulaRule(formula=['$A2<>"成功"'], fill=missing_fill),
        )

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(max(max_length + 2, 10), 32)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(xlsx_path)
    return csv_path, xlsx_path


def query_departure_city(city, depart_date, direct=False, timeout=20):
    departure_code = city_api.fetch_city_code(city, timeout=timeout)
    prices = lowest_price_api.fetch_lowest_price_calendar(
        departure_code,
        DESTINATION_CODE,
        direct=direct,
        timeout=timeout,
    )
    price = prices.get(depart_date)
    if price is None:
        return {
            "状态": "接口无价格",
            "出发城市": city,
            "到达城市": DESTINATION_CITY,
            "出发日期": depart_date,
            "最低价": "",
            "返回日期数": len(prices),
            "价格说明": "lowestPrice API 未返回该日期价格",
        }
    return {
        "状态": "成功",
        "出发城市": city,
        "到达城市": DESTINATION_CITY,
        "出发日期": depart_date,
        "最低价": price,
        "返回日期数": len(prices),
        "价格说明": "lowestPrice API 单程低价日历价格",
    }


def sort_rows(rows):
    frame = pd.DataFrame(rows)
    frame["_sort_price"] = pd.to_numeric(frame["最低价"], errors="coerce")
    frame = frame.sort_values("_sort_price", ascending=True, na_position="last")
    return frame.drop(columns=["_sort_price"])


def sort_round_trip_rows(rows):
    frame = pd.DataFrame(rows)
    frame["_sort_price"] = pd.to_numeric(frame["总价"], errors="coerce")
    frame = frame.sort_values("_sort_price", ascending=True, na_position="last")
    return frame.drop(columns=["_sort_price"])


def run_job(cities, depart_date, run_day, direct=False, timeout=20, request_interval=2.0, include_xinjiang_local=False):
    rows = []
    selected_cities = filter_cities(cities, include_xinjiang_local=include_xinjiang_local)
    log(f"国内 API 任务准备完成：出发城市 {len(selected_cities)} 个，目的地 {DESTINATION_CITY}，出发日期 {depart_date}")
    for index, city in enumerate(selected_cities, start=1):
        log(f"开始查询 {index}/{len(selected_cities)}：{city}->{DESTINATION_CITY} {depart_date}")
        try:
            row = query_departure_city(
                city,
                depart_date,
                direct=direct,
                timeout=timeout,
            )
        except Exception as e:
            row = {
                "状态": "查询异常",
                "出发城市": city,
                "到达城市": DESTINATION_CITY,
                "出发日期": depart_date,
                "最低价": "",
                "返回日期数": "",
                "价格说明": f"{type(e).__name__}: {e}",
            }
        rows.append(row)
        log(f"完成查询 {index}/{len(selected_cities)}：{city}->{DESTINATION_CITY} {row['状态']} {row['最低价']}")
        if index < len(selected_cities) and request_interval > 0:
            time.sleep(request_interval)

    frame = sort_rows(rows)
    output_file = result_file_path(run_day, depart_date)
    csv_path, xlsx_path = write_table_outputs(frame, output_file, highlight_missing=True)
    log(f"国内 API 结果已生成：{csv_path}")
    log(f"国内 API Excel 已生成：{xlsx_path}")
    return csv_path, xlsx_path


def build_round_trip_rows(outbound_city, return_city, outbound_prices, inbound_prices, depart_dates, stay_days):
    rows = []
    for depart_date in depart_dates:
        return_date = (
            dt.strptime(depart_date, "%Y-%m-%d")
            + timedelta(days=stay_days)
        ).strftime("%Y-%m-%d")
        outbound_price = outbound_prices.get(depart_date)
        inbound_price = inbound_prices.get(return_date)
        if outbound_price is not None and inbound_price is not None:
            status = "成功"
            total_price = outbound_price + inbound_price
            price_note = "lowestPrice API 两段单程低价相加"
        else:
            missing = []
            if outbound_price is None:
                missing.append("去程无价格")
            if inbound_price is None:
                missing.append("返程无价格")
            status = "；".join(missing)
            total_price = ""
            price_note = "lowestPrice API 未返回完整往返日期价格"
        rows.append(
            {
                "状态": status,
                "去程出发城市": outbound_city,
                "目的城市": DESTINATION_CITY,
                "返程到达城市": return_city,
                "去程出发日期": depart_date,
                "返程出发日期": return_date,
                "行程天数": stay_days,
                "总价": total_price,
                "去程最低价": outbound_price if outbound_price is not None else "",
                "返程最低价": inbound_price if inbound_price is not None else "",
                "去程返回日期数": len(outbound_prices),
                "返程返回日期数": len(inbound_prices),
                "价格说明": price_note,
            }
        )
    return rows


def fetch_route_calendar_cached(cache, departure_code, arrival_code, direct=False, timeout=20, request_interval=2.0):
    route = (departure_code, arrival_code)
    if route in cache:
        return cache[route]
    cache[route] = lowest_price_api.fetch_lowest_price_calendar(
        departure_code,
        arrival_code,
        direct=direct,
        timeout=timeout,
    )
    if request_interval > 0:
        time.sleep(request_interval)
    return cache[route]


def run_round_trip_job(
    cities,
    return_cities,
    month,
    stay_days,
    run_day,
    open_jaw=False,
    include_xinjiang_local=False,
    direct=False,
    timeout=20,
    request_interval=2.0,
):
    rows = []
    depart_dates = month_dates(month)
    selected_cities = filter_cities(cities, include_xinjiang_local=include_xinjiang_local)
    selected_return_cities = filter_cities(return_cities, include_xinjiang_local=include_xinjiang_local)
    pairs = []
    if open_jaw:
        for outbound_city in selected_cities:
            for return_city in selected_return_cities:
                if outbound_city != return_city:
                    pairs.append((outbound_city, return_city))
    else:
        pairs = [(city, city) for city in selected_cities]

    log(
        f"国内往返 API 任务准备完成：去程出发城市 {len(selected_cities)} 个，"
        f"返程到达城市 {len(selected_return_cities)} 个，组合 {len(pairs)} 组，"
        f"目的地 {DESTINATION_CITY}，去程月份 {month}，行程天数 {stay_days}"
    )
    city_codes = {}
    calendars = {}
    cities_to_resolve = sorted(set(selected_cities + selected_return_cities))
    for index, city in enumerate(cities_to_resolve, start=1):
        log(f"解析城市 {index}/{len(cities_to_resolve)}：{city}")
        city_codes[city] = city_api.fetch_city_code(city, timeout=timeout)
        if index < len(cities_to_resolve) and request_interval > 0:
            time.sleep(request_interval)

    for index, (outbound_city, return_city) in enumerate(pairs, start=1):
        log(f"开始组合 {index}/{len(pairs)}：{outbound_city}->{DESTINATION_CITY}->{return_city} {month} {stay_days}天")
        try:
            outbound_prices = fetch_route_calendar_cached(
                calendars,
                city_codes[outbound_city],
                DESTINATION_CODE,
                direct=direct,
                timeout=timeout,
                request_interval=request_interval,
            )
            inbound_prices = fetch_route_calendar_cached(
                calendars,
                DESTINATION_CODE,
                city_codes[return_city],
                direct=direct,
                timeout=timeout,
                request_interval=request_interval,
            )
            city_rows = build_round_trip_rows(
                outbound_city,
                return_city,
                outbound_prices,
                inbound_prices,
                depart_dates,
                stay_days,
            )
        except Exception as e:
            city_rows = [
                {
                    "状态": "查询异常",
                    "去程出发城市": outbound_city,
                    "目的城市": DESTINATION_CITY,
                    "返程到达城市": return_city,
                    "去程出发日期": depart_date,
                    "返程出发日期": (
                        dt.strptime(depart_date, "%Y-%m-%d")
                        + timedelta(days=stay_days)
                    ).strftime("%Y-%m-%d"),
                    "行程天数": stay_days,
                    "总价": "",
                    "去程最低价": "",
                    "返程最低价": "",
                    "去程返回日期数": "",
                    "返程返回日期数": "",
                    "价格说明": f"{type(e).__name__}: {e}",
                }
                for depart_date in depart_dates
            ]
        rows.extend(city_rows)
        success_count = sum(1 for row in city_rows if row["状态"] == "成功")
        log(f"完成组合 {index}/{len(pairs)}：{outbound_city}->{DESTINATION_CITY}->{return_city} 成功日期 {success_count}/{len(city_rows)}")

    frame = sort_round_trip_rows(rows)
    output_file = (
        open_jaw_result_file_path(run_day, month, stay_days)
        if open_jaw
        else round_trip_result_file_path(run_day, month, stay_days)
    )
    csv_path, xlsx_path = write_table_outputs(frame, output_file, highlight_missing=True)
    log(f"国内往返 API 结果已生成：{csv_path}")
    log(f"国内往返 API Excel 已生成：{xlsx_path}")
    return csv_path, xlsx_path


def parse_args():
    parser = argparse.ArgumentParser(description="通过携程 lowestPrice API 查询全国各地往返乌鲁木齐低价")
    parser.add_argument("--round-trip", action="store_true", help="查询往返价格")
    parser.add_argument("--open-jaw", action="store_true", help="查询异地返程组合，例如上海->乌鲁木齐->杭州")
    parser.add_argument("--month", default="2026-06", help="往返查询的去程月份，格式 YYYY-MM")
    parser.add_argument("--stay-days", type=int, default=10, help="返程日期 = 去程日期 + stay-days")
    parser.add_argument("--date", default="2026-05-25", help="出发日期，格式 YYYY-MM-DD")
    parser.add_argument("--cities", nargs="+", default=None, help="出发城市列表")
    parser.add_argument("--return-cities", nargs="+", default=None, help="返程到达城市列表，不填则复用出发城市列表")
    parser.add_argument("--jiangzhehuwan", action="store_true", help="使用江浙沪皖机场城市作为出发和返程城市范围")
    parser.add_argument("--include-xinjiang-local", action="store_true", help="包含新疆本地城市")
    parser.add_argument("--run-day", default=dt.now().strftime("%Y-%m-%d"), help="结果输出日期目录")
    parser.add_argument("--direct", action="store_true", help="只查直飞低价")
    parser.add_argument("--timeout", type=int, default=20, help="单个 API 请求超时秒数")
    parser.add_argument("--request-interval", type=float, default=2.0, help="每次 API 请求之间的等待秒数")
    return parser.parse_args()


def main():
    args = parse_args()
    cities = (
        JIANGZHEHUWAN_AIRPORT_CITIES
        if args.jiangzhehuwan
        else args.cities or DEFAULT_DEPARTURE_CITIES
    )
    return_cities = args.return_cities or cities
    if args.round_trip:
        output_file = run_round_trip_job(
            cities,
            return_cities,
            args.month,
            args.stay_days,
            args.run_day,
            open_jaw=args.open_jaw,
            include_xinjiang_local=args.include_xinjiang_local,
            direct=args.direct,
            timeout=args.timeout,
            request_interval=args.request_interval,
        )
    else:
        output_file = run_job(
            cities,
            args.date,
            args.run_day,
            direct=args.direct,
            timeout=args.timeout,
            request_interval=args.request_interval,
            include_xinjiang_local=args.include_xinjiang_local,
        )
    log(f"任务结束：{output_file}")


if __name__ == "__main__":
    main()
